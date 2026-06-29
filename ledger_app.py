import sys
import os
import pandas as pd
from datetime import date
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QListWidget, QListWidgetItem, QPushButton, QTableWidget, 
                             QTableWidgetItem, QSplitter, QDialog, 
                             QFormLayout, QLineEdit, QDateEdit, QHeaderView, QLabel, QMessageBox,
                             QFileDialog, QDialogButtonBox, QComboBox)
from PyQt6.QtCore import Qt, QRegularExpression, QDate, QStringListModel, QEvent
from PyQt6.QtWidgets import QCompleter, QMenu
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtGui import QTextDocument, QPageSize, QColor, QRegularExpressionValidator

def indian_format(n):
    """Format a number in the Indian numbering system (e.g. 1,23,456.78)."""
    s = f"{n:.2f}"
    sign = ''
    if s.startswith('-'):
        sign = '-'
        s = s[1:]
    int_part, dec_part = s.split('.')
    dec_part = dec_part.rstrip('0')
    dec_part = f".{dec_part}" if dec_part else ''
    if len(int_part) > 3:
        last3 = int_part[-3:]
        rest = int_part[:-3]
        groups = []
        while rest:
            groups.append(rest[-2:])
            rest = rest[:-2]
        groups.reverse()
        return sign + ','.join(groups) + ',' + last3 + dec_part
    return sign + int_part + dec_part

class CreatePartyDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("New Party")
        self.resize(420, 200)
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(28, 28, 28, 28)

        heading = QLabel("Create a New Party")
        heading.setStyleSheet("font-size: 18px; font-weight: 700; color: #0f172a;")
        layout.addWidget(heading)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Enter party name...")
        layout.addWidget(self.name_input)

        self.btn = QPushButton("  + Create Party")
        self.btn.setObjectName("btn_dialog_add")
        self.btn.clicked.connect(self.accept)
        layout.addWidget(self.btn)
        layout.addStretch()

    def accept(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Validation Error", "Party name cannot be empty.")
            return
        super().accept()

    def get_name(self):
        return self.name_input.text().strip()


class TransactionDialog(QDialog):
    def __init__(self, parties, parent=None):
        super().__init__(parent)
        self.parties = parties
        self.setWindowTitle("New Transaction")
        self.resize(520, 380)
        layout = QFormLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(24, 24, 24, 24)

        self.date_input = QLineEdit()
        self.date_input.setInputMask("00-00-0000;_")
        self.date_input.setText(date.today().strftime("%d-%m-%Y"))
        self.date_input.setPlaceholderText("DD-MM-YYYY")
        self.party_input = QLineEdit()
        self.party_input.setPlaceholderText("Type to search party...")
        completer = QCompleter(parties)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.party_input.setCompleter(completer)

        self.desc_input = QLineEdit()
        self.status_input = QComboBox()
        self.status_input.addItems(["To Receive", "To Pay"])
        self.qty_input = QLineEdit()
        self.rate_input = QLineEdit()
        self.total_input = QLineEdit()
        self.total_input.setPlaceholderText("Auto or enter directly")

        self.qty_input.textChanged.connect(self.auto_calc_total)
        self.rate_input.textChanged.connect(self.auto_calc_total)

        layout.addRow("Date:", self.date_input)
        layout.addRow("Party Name:", self.party_input)
        layout.addRow("Description:", self.desc_input)
        layout.addRow("Status:", self.status_input)
        layout.addRow("Quantity:", self.qty_input)
        layout.addRow("Rate:", self.rate_input)
        layout.addRow("Total:", self.total_input)

        self.btn = QPushButton("Save Transaction")
        self.btn.setObjectName("btn_save")
        self.btn.clicked.connect(self.accept)
        layout.addRow(self.btn)

    def auto_calc_total(self):
        qty_str = self.qty_input.text().strip()
        rate_str = self.rate_input.text().strip()
        if qty_str and rate_str:
            try:
                qty = float(qty_str)
                rate = float(rate_str)
                self.total_input.setText(str(qty * rate))
            except ValueError:
                pass

    def accept(self):
        date_str = self.date_input.text().strip()
        party_str = self.party_input.text().strip()
        qty_str = self.qty_input.text().strip()
        rate_str = self.rate_input.text().strip()
        total_str = self.total_input.text().strip()

        if not date_str:
            QMessageBox.warning(self, "Validation Error", "Date is required.")
            return
        if not party_str:
            QMessageBox.warning(self, "Validation Error", "Please create a party first.")
            return
        if party_str not in self.parties:
            QMessageBox.warning(self, "Validation Error", f"Party '{party_str}' does not exist. Select a valid party.")
            return
        if not self.date_input.hasAcceptableInput():
            QMessageBox.warning(self, "Validation Error", "Please enter a complete date (DD-MM-YYYY).")
            return

        qty_valid = False
        rate_valid = False
        total_valid = False

        if qty_str:
            try:
                qty = float(qty_str)
                if qty >= 0:
                    qty_valid = True
                else:
                    QMessageBox.warning(self, "Validation Error", "Quantity cannot be negative.")
                    return
            except ValueError:
                QMessageBox.warning(self, "Validation Error", "Quantity must be a valid number.")
                return

        if rate_str:
            try:
                rate = float(rate_str)
                if rate >= 0:
                    rate_valid = True
                else:
                    QMessageBox.warning(self, "Validation Error", "Rate cannot be negative.")
                    return
            except ValueError:
                QMessageBox.warning(self, "Validation Error", "Rate must be a valid number.")
                return

        if total_str:
            try:
                total = float(total_str)
                if total >= 0:
                    total_valid = True
                else:
                    QMessageBox.warning(self, "Validation Error", "Total cannot be negative.")
                    return
            except ValueError:
                QMessageBox.warning(self, "Validation Error", "Total must be a valid number.")
                return

        if not ((qty_valid and rate_valid) or total_valid):
            QMessageBox.warning(self, "Validation Error",
                "Enter both Quantity & Rate, or enter Total directly.")
            return

        super().accept()

    def get_data(self):
        qty_str = self.qty_input.text().strip()
        rate_str = self.rate_input.text().strip()
        total_str = self.total_input.text().strip()

        qty = float(qty_str) if qty_str else 0.0
        rate = float(rate_str) if rate_str else 0.0
        total = float(total_str) if total_str else (qty * rate)

        return {
            "Date": self.date_input.text().strip(),
            "Party Name": self.party_input.text().strip(),
            "Description": self.desc_input.text().strip(),
            "Status": self.status_input.currentText(),
            "Quantity": qty,
            "Rate": rate,
            "Total": total
        }

class PartySummaryDialog(QDialog):
    def __init__(self, file_path, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.setWindowTitle("Party Summary")
        self.resize(550, 500)
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        heading = QLabel("Party-wise Grand Total")
        heading.setStyleSheet("font-size: 20px; font-weight: 700; color: #0f172a;")
        layout.addWidget(heading)

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Party Name", "Grand Total"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #e2e8f0; border-radius: 8px; font-size: 14px; }
            QHeaderView::section { background-color: #f1f5f9; padding: 10px 8px;
                font-weight: 600; font-size: 13px; color: #475569; border: none;
                border-bottom: 1px solid #e2e8f0; }
        """)
        layout.addWidget(self.table)

        self.total_label = QLabel()
        self.total_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.total_label.setStyleSheet("font-size: 16px; font-weight: 700; color: #0f172a; padding: 8px 12px;")
        layout.addWidget(self.total_label)

        btn_close = QPushButton("Close")
        btn_close.setStyleSheet("padding: 10px 24px; font-size: 14px; font-weight: 600;")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close, alignment=Qt.AlignmentFlag.AlignRight)

        self.load_data()

    def load_data(self):
        try:
            xl = pd.ExcelFile(self.file_path)
            parties = [s for s in xl.sheet_names if s != 'Sheet1']
            rows = []
            overall = 0.0
            for party in parties:
                df = pd.read_excel(self.file_path, sheet_name=party)
                if df.empty:
                    continue
                df['Total'] = pd.to_numeric(df.get('Total', 0), errors='coerce').fillna(0)
                credit = df[df['Status'] == 'To Receive']['Total'].sum()
                debit = df[df['Status'] == 'To Pay']['Total'].sum()
                gt = credit - debit
                rows.append((party, gt))
                overall += gt
            self.table.setRowCount(len(rows))
            for i, (party, gt) in enumerate(rows):
                name_item = QTableWidgetItem(party)
                name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(i, 0, name_item)
                total_item = QTableWidgetItem(indian_format(gt))
                total_item.setFlags(total_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                font = total_item.font()
                font.setWeight(600)
                total_item.setFont(font)
                self.table.setItem(i, 1, total_item)
            self.total_label.setText(f"Overall Grand Total: {indian_format(overall)}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not load summary: {e}")

class LedgerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_path = "ledger.xlsx"
        self._current_party = None
        self._status_filter = None
        self._date_from = None
        self._date_to = None
        self._blocked_descriptions = set()
        self._current_suggestion = None
        self.init_data_file()
        self.init_ui()

    SHEET_COLUMNS = ['Date', 'Description', 'Quantity', 'Rate', 'Status', 'Total']

    def init_data_file(self):
        if not os.path.exists(self.file_path):
            pd.DataFrame(columns=self.SHEET_COLUMNS).to_excel(self.file_path, sheet_name='Sheet1', index=False)
        blocked_path = os.path.join(os.path.dirname(os.path.abspath(self.file_path)), 'blocked_descriptions.json')
        if os.path.exists(blocked_path):
            import json
            with open(blocked_path, 'r') as f:
                self._blocked_descriptions = set(json.load(f))
        else:
            self._blocked_descriptions = set()

    def load_all_data(self):
        try:
            xl = pd.ExcelFile(self.file_path)
            all_data = []
            for sheet in xl.sheet_names:
                if sheet == 'Sheet1':
                    continue
                df = pd.read_excel(self.file_path, sheet_name=sheet)
                if not df.empty:
                    df['Party Name'] = sheet
                    df['_sheet_name'] = sheet
                    df['_sheet_row'] = df.index
                    # Ensure columns exist (backward compat)
                    if 'Total' not in df.columns:
                        df['Total'] = 0.0
                    if 'Status' not in df.columns:
                        df['Status'] = 'Credit'
                    all_data.append(df)
            cols = ['Date', 'Party Name', 'Description', 'Status', 'Quantity', 'Rate', 'Total', '_sheet_name', '_sheet_row']
            return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame(columns=cols)
        except Exception as e:
            return pd.DataFrame(columns=['Date', 'Party Name', 'Description', 'Status', 'Quantity', 'Rate', 'Total', '_sheet_name', '_sheet_row'])

    def _save_blocked_descriptions(self):
        import json
        blocked_path = os.path.join(os.path.dirname(os.path.abspath(self.file_path)), 'blocked_descriptions.json')
        with open(blocked_path, 'w') as f:
            json.dump(list(self._blocked_descriptions), f)

    def _remove_desc_suggestion(self, desc):
        self._blocked_descriptions.add(desc)
        self._save_blocked_descriptions()
        # Update the completer model
        if hasattr(self, '_desc_completer_model'):
            all_descs = self._get_all_descriptions()
            self._desc_completer_model.setStringList(all_descs)

    def _get_all_descriptions(self):
        all_data = self.load_all_data()
        all_descs = sorted(set(all_data['Description'].dropna().astype(str).tolist()))
        return [d for d in all_descs if d not in self._blocked_descriptions]

    def init_ui(self):
        self.setWindowTitle("Party Ledger Manager")
        self.resize(1300, 800)

        STYLE = """
            QMainWindow, QWidget {
                background-color: #f1f5f9;
            }
            QLabel {
                font-size: 14px;
                color: #334155;
            }
            QLabel#sidebar_title {
                font-size: 13px;
                font-weight: 700;
                color: #64748b;
                text-transform: uppercase;
                letter-spacing: 1px;
                padding: 0 4px;
                margin-bottom: 4px;
            }
            QLabel#party_count {
                font-size: 12px;
                color: #94a3b8;
                padding: 0 4px;
                margin-bottom: 6px;
            }
            QSplitter::handle {
                background-color: #e2e8f0;
                width: 1px;
            }
            QDialog {
                background-color: #ffffff;
                border-radius: 12px;
            }
            QLineEdit {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 10px 14px;
                font-size: 14px;
                color: #1e293b;
                outline: none;
                selection-background-color: #6366f1;
                selection-color: white;
            }
            QLineEdit:focus {
                border-color: #6366f1;
                border-width: 2px;
                padding: 9px 13px;
            }
            QLineEdit#inp_inline {
                border: none;
                background: transparent;
                padding: 4px 8px;
                font-size: 14px;
                color: #1e293b;
            }
            QLineEdit#inp_inline:focus {
                border: none;
                background: transparent;
            }
            QComboBox {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                color: #1e293b;
                outline: none;
            }
            QComboBox:focus {
                border-color: #6366f1;
            }
            QComboBox::drop-down {
                border: none;
                padding-right: 8px;
            }
            QComboBox::down-arrow {
                image: none;
                width: 0;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 4px;
                color: #1e293b;
                font-size: 14px;
                outline: none;
                selection-background-color: #f1f5f9;
                selection-color: #6366f1;
            }
            QComboBox QAbstractItemView::item {
                padding: 8px 12px;
                border-radius: 6px;
                min-height: 32px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #f8fafc;
            }
            QPushButton {
                background-color: #6366f1;
                color: white;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: 600;
                font-size: 14px;
                border: none;
                outline: none;
            }
            QPushButton:hover {
                background-color: #4f46e5;
            }
            QPushButton:pressed {
                background-color: #4338ca;
            }
            QPushButton:focus {
                outline: none;
            }
            QPushButton#btn_create {
                background-color: #6366f1;
                font-size: 14px;
                padding: 10px 16px;
                border-radius: 8px;
            }
            QPushButton#btn_create:hover {
                background-color: #4f46e5;
            }
            QPushButton#btn_remove_party {
                background-color: #ffffff;
                color: #ef4444;
                border: 1px solid #fca5a5;
                font-size: 14px;
                padding: 10px 16px;
                border-radius: 8px;
            }
            QPushButton#btn_remove_party:hover {
                background-color: #fef2f2;
                border-color: #ef4444;
            }
            QPushButton#btn_remove_party:pressed {
                background-color: #fee2e2;
            }
            QPushButton#btn_print {
                background-color: #059669;
                padding: 10px 24px;
                border-radius: 8px;
            }
            QPushButton#btn_print:hover {
                background-color: #047857;
            }
            QPushButton#btn_print:pressed {
                background-color: #065f46;
            }
            QPushButton#btn_date_filter {
                background-color: #ffffff;
                color: #475569;
                border: 1px solid #cbd5e1;
                padding: 10px 24px;
                border-radius: 8px;
                font-weight: 500;
            }
            QPushButton#btn_date_filter:hover {
                background-color: #f8fafc;
                border-color: #94a3b8;
            }
            QPushButton#btn_date_filter:pressed {
                background-color: #f1f5f9;
            }
            QPushButton#btn_summary {
                background-color: #4f46e5;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
                font-size: 13px;
                font-weight: 600;
            }
            QPushButton#btn_summary:hover {
                background-color: #4338ca;
            }
            QPushButton#btn_active_filter {
                background-color: #eef2ff;
                color: #6366f1;
                border: 1px solid #a5b4fc;
                padding: 10px 24px;
                border-radius: 8px;
                font-weight: 500;
            }
            QPushButton#btn_active_filter:hover {
                background-color: #e0e7ff;
            }
            QPushButton#btn_remove {
                background-color: transparent;
                color: #94a3b8;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 0;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton#btn_remove:hover {
                background-color: #fef2f2;
                border-color: #f87171;
                color: #ef4444;
            }
            QPushButton#btn_remove:pressed {
                background-color: #fee2e2;
            }
            QPushButton#btn_save {
                background-color: #6366f1;
                font-size: 15px;
                padding: 12px 24px;
                border-radius: 8px;
                min-height: 20px;
            }
            QPushButton#btn_save:hover {
                background-color: #4f46e5;
            }
            QPushButton#btn_dialog_add {
                background-color: #6366f1;
                font-size: 14px;
                padding: 10px 24px;
                border-radius: 8px;
            }
            QListWidget {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 6px;
                font-size: 14px;
                outline: none;
            }
            QListWidget:focus {
                outline: none;
            }
            QListWidget::item {
                padding: 10px 14px;
                border-radius: 8px;
                color: #334155;
                font-weight: 500;
                margin: 2px 0;
            }
            QListWidget::item:hover {
                background-color: #f1f5f9;
                color: #1e293b;
            }
            QListWidget::item:selected {
                background-color: #00008B;
                color: #ffffff;
                font-weight: 600;
            }
            QTableWidget {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                gridline-color: #f1f5f9;
                font-size: 14px;
                color: #1e293b;
                outline: none;
            }
            QTableWidget:focus {
                outline: none;
            }
            QTableWidget::item {
                padding: 4px 8px;
                border-bottom: 1px solid #f1f5f9;
            }
            QTableWidget::item:selected {
                background-color: #eef2ff;
                color: #6366f1;
            }
            QHeaderView::section {
                background-color: #f8fafc;
                padding: 12px 8px;
                font-weight: 600;
                font-size: 13px;
                color: #64748b;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                border-right: 1px solid #f1f5f9;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            QScrollBar:vertical {
                background: #f1f5f9;
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #cbd5e1;
                border-radius: 4px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: #94a3b8;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0;
            }
            QScrollBar:horizontal {
                background: #f1f5f9;
                height: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:horizontal {
                background: #cbd5e1;
                border-radius: 4px;
                min-width: 30px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #94a3b8;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0;
            }
            QAbstractItemView {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 6px;
                color: #1e293b;
                font-size: 13px;
                outline: none;
            }
            QAbstractItemView:focus {
                outline: none;
            }
            QAbstractItemView::item {
                padding: 8px 12px;
                border-radius: 6px;
            }
            QAbstractItemView::item:hover {
                background-color: #f1f5f9;
            }
            QAbstractItemView::item:selected {
                background-color: #00008B;
                color: #ffffff;
            }
        """
        self.setStyleSheet(STYLE)

        main_widget = QWidget()
        layout = QHBoxLayout(main_widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Sidebar
        sidebar = QVBoxLayout()
        sidebar.setContentsMargins(16, 16, 12, 16)
        sidebar.setSpacing(8)

        title_label = QLabel("PARTIES")
        title_label.setObjectName("sidebar_title")
        sidebar.addWidget(title_label)

        self.party_count_label = QLabel("")
        self.party_count_label.setObjectName("party_count")
        sidebar.addWidget(self.party_count_label)

        # Status filter buttons
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(4)
        self._btn_filter_all = QPushButton("All")
        self._btn_filter_all.setCheckable(True)
        self._btn_filter_all.setChecked(True)
        self._btn_filter_receive = QPushButton("To Receive")
        self._btn_filter_receive.setCheckable(True)
        self._btn_filter_pay = QPushButton("To Pay")
        self._btn_filter_pay.setCheckable(True)
        for btn in [self._btn_filter_all, self._btn_filter_receive, self._btn_filter_pay]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #ffffff; color: #64748b; border: 1px solid #e2e8f0;
                    border-radius: 6px; padding: 6px 10px; font-size: 11px; font-weight: 500;
                    outline: none;
                }
                QPushButton:hover {
                    background-color: #f1f5f9; border-color: #cbd5e1;
                }
                QPushButton:checked {
                    background-color: #eef2ff; color: #6366f1; border-color: #a5b4fc;
                }
            """)
        self._btn_filter_all.clicked.connect(lambda: self._set_status_filter(None))
        self._btn_filter_receive.clicked.connect(lambda: self._set_status_filter('To Receive'))
        self._btn_filter_pay.clicked.connect(lambda: self._set_status_filter('To Pay'))
        filter_layout.addWidget(self._btn_filter_all)
        filter_layout.addWidget(self._btn_filter_receive)
        filter_layout.addWidget(self._btn_filter_pay)
        sidebar.addLayout(filter_layout)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search parties...")
        self.search_input.textChanged.connect(self.filter_party_list)
        sidebar.addWidget(self.search_input)

        self.party_list = QListWidget()
        self.party_list.itemClicked.connect(self.filter_by_party)
        self.party_list.itemChanged.connect(self._rename_party)

        btn_add_party = QPushButton("  + New Party")
        btn_add_party.setObjectName("btn_create")
        btn_add_party.clicked.connect(self.add_party)
        btn_remove_party = QPushButton("  − Remove")
        btn_remove_party.setObjectName("btn_remove_party")
        btn_remove_party.clicked.connect(self.remove_party)

        sidebar.addWidget(btn_add_party)
        sidebar.addWidget(btn_remove_party)
        sidebar.addWidget(self.party_list)
        
        # Main Area
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(48)
        self.table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.table.setWordWrap(True)
        self.table.cellChanged.connect(self._on_cell_changed)
        self.table.cellDoubleClicked.connect(self._on_cell_double_clicked)

        btn_print = QPushButton("  Export Report")
        btn_print.setObjectName("btn_print")
        btn_print.clicked.connect(self.print_report)

        self.btn_date_filter = QPushButton("  Date Filter")
        self.btn_date_filter.setObjectName("btn_date_filter")
        self.btn_date_filter.clicked.connect(self._apply_date_filter)

        btn_summary = QPushButton("  Party Summary")
        btn_summary.setObjectName("btn_summary")
        btn_summary.clicked.connect(self.show_party_summary)
        
        main_area = QVBoxLayout()
        main_area.setContentsMargins(12, 16, 16, 16)
        main_area.setSpacing(12)

        header_row = QHBoxLayout()
        header_row.setSpacing(0)
        self._header_title = QLabel("All Transactions")
        self._header_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0f172a;")
        header_row.addWidget(self._header_title)
        header_row.addStretch()
        
        top_bar = QHBoxLayout()
        top_bar.setSpacing(10)
        top_bar.addWidget(btn_summary)
        top_bar.addWidget(self.btn_date_filter)
        top_bar.addWidget(btn_print)

        main_area.addLayout(header_row)
        main_area.addLayout(top_bar)
        main_area.addWidget(self.table)
        
        # Layout Assembly
        splitter = QSplitter(Qt.Orientation.Horizontal)
        side_widget = QWidget()
        side_widget.setLayout(sidebar)
        main_widget_area = QWidget()
        main_widget_area.setLayout(main_area)
        
        splitter.addWidget(side_widget)
        splitter.addWidget(main_widget_area)
        splitter.setSizes([300, 1000])  # Initial column split width
        
        layout.addWidget(splitter)
        self.setCentralWidget(main_widget)
        self.refresh_view()

    def refresh_view(self, df=None):
        if df is None:
            self._current_party = None
            df = self.load_all_data()
        
        # Always build a clean DataFrame with the 6 columns in fixed order.
        cols = ['Date', 'Party Name', 'Description', 'Status', 'Quantity', 'Rate', 'Total']

        if not df.empty:
            meta_cols = ['_sheet_name', '_sheet_row']
            meta = df[meta_cols].copy() if all(c in df.columns for c in meta_cols) else None

            for c in ['Quantity', 'Rate', 'Total']:
                if c not in df.columns:
                    df[c] = 0.0
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
            if 'Status' not in df.columns:
                df['Status'] = 'Credit'
            df['Status'] = df['Status'].astype(str)
            calc_total = df['Quantity'] * df['Rate']
            mask = (df['Total'] == 0) & (calc_total != 0)
            df.loc[mask, 'Total'] = calc_total[mask]
            df['Party Name'] = df.get('Party Name', '')
            df['Date'] = df.get('Date', '').astype(str)
            df['Description'] = df.get('Description', '').astype(str)

            df['_date_parsed'] = pd.to_datetime(df['Date'], format='%d-%m-%Y', errors='coerce')

            if self._date_from and self._date_to:
                from_dt = pd.to_datetime(self._date_from, format='%d-%m-%Y', errors='coerce')
                to_dt = pd.to_datetime(self._date_to, format='%d-%m-%Y', errors='coerce')
                if pd.notna(from_dt) and pd.notna(to_dt):
                    mask = (df['_date_parsed'] >= from_dt) & (df['_date_parsed'] <= to_dt)
                    df = df[mask]
                    if meta is not None:
                        meta = meta.loc[df.index]
            if self._status_filter:
                mask = df['Status'] == self._status_filter
                df = df[mask]
                if meta is not None:
                    meta = meta.loc[df.index]
            sort_idx = df.sort_values(by='_date_parsed', ascending=False).index
            df = df.loc[sort_idx]
            if meta is not None:
                meta = meta.loc[sort_idx]
            clean = df.reindex(columns=cols).fillna('')
        else:
            clean = pd.DataFrame(columns=cols)
            meta = None

        self.table.blockSignals(True)
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self.table.setColumnCount(8)
        headers = cols + ['Action']
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setColumnWidth(0, 120)
        self.table.setColumnWidth(1, 150)
        self.table.setColumnWidth(2, 300)
        self.table.setColumnWidth(3, 110)
        self.table.setColumnWidth(4, 70)
        self.table.setColumnWidth(5, 70)
        self.table.setColumnWidth(6, 130)
        self.table.setColumnWidth(7, 50)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        show_input = self._current_party is not None
        show_party = self._current_party is None
        self.table.setColumnHidden(1, not show_party)
        clean['Total'] = pd.to_numeric(clean['Total'], errors='coerce').fillna(0.0)
        credit_mask = clean['Status'] == 'To Receive'
        grand_total = clean.loc[credit_mask, 'Total'].sum() - clean.loc[~credit_mask, 'Total'].sum()

        if show_input:
            self.table.setRowCount(len(clean) + 2)  # input + data + total
            inp_style = "border: none; background: transparent; padding: 4px 8px; font-size: 14px;"
            row_h = 48
            self.table.setRowHeight(0, row_h)
            # Row 0: inline input row
            self._inp_date = QLineEdit()
            self._inp_date.setObjectName("inp_inline")
            self._inp_date.setInputMask("00-00-0000;_")
            self._inp_date.setText(date.today().strftime("%d-%m-%Y"))
            self._inp_date.setPlaceholderText("DD-MM-YYYY")
            self._inp_date.returnPressed.connect(self._add_inline_transaction)
            self.table.setCellWidget(0, 0, self._inp_date)

            party_label = QLabel(self._current_party)
            party_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            party_label.setStyleSheet("font-weight: 600; color: #6366f1; border: none; background: transparent;")
            self.table.setCellWidget(0, 1, party_label)

            self._inp_desc = QLineEdit()
            self._inp_desc.setObjectName("inp_inline")
            self._inp_desc.setPlaceholderText("Description")
            self._inp_desc.returnPressed.connect(self._add_inline_transaction)
            # Description autocomplete from all existing entries
            all_descs = self._get_all_descriptions()
            self._desc_completer_model = QStringListModel(all_descs)
            desc_completer = QCompleter(self._desc_completer_model)
            desc_completer.setFilterMode(Qt.MatchFlag.MatchContains)
            desc_completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
            desc_completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            self._inp_desc.setCompleter(desc_completer)
            # Track current suggestion for removal via Delete key
            desc_completer.highlighted.connect(self._on_suggestion_highlighted)
            desc_completer.activated.connect(self._on_desc_selected)
            self._desc_completer = desc_completer
            self._inp_desc.installEventFilter(self)
            # Right-click removal on popup
            popup = desc_completer.popup()
            popup.viewport().installEventFilter(self)
            self._desc_popup_viewport = popup.viewport()
            self.table.setCellWidget(0, 2, self._inp_desc)

            self._inp_status = QComboBox()
            self._inp_status.addItems(["To Receive", "To Pay"])
            self._inp_status.setStyleSheet("border: none; background: transparent; padding: 4px 8px; font-size: 14px; color: #1e293b;")
            self.table.setCellWidget(0, 3, self._inp_status)

            num_validator = QRegularExpressionValidator(QRegularExpression(r'^\d*\.?\d*$'))

            self._inp_qty = QLineEdit()
            self._inp_qty.setObjectName("inp_inline")
            self._inp_qty.setPlaceholderText("Qty")
            self._inp_qty.setValidator(num_validator)
            self._inp_qty.textChanged.connect(self._inline_auto_calc)
            self._inp_qty.returnPressed.connect(self._add_inline_transaction)
            self.table.setCellWidget(0, 4, self._inp_qty)

            self._inp_rate = QLineEdit()
            self._inp_rate.setObjectName("inp_inline")
            self._inp_rate.setPlaceholderText("Rate")
            self._inp_rate.setValidator(num_validator)
            self._inp_rate.textChanged.connect(self._inline_auto_calc)
            self._inp_rate.returnPressed.connect(self._add_inline_transaction)
            self.table.setCellWidget(0, 5, self._inp_rate)

            self._inp_total = QLineEdit()
            self._inp_total.setObjectName("inp_inline")
            self._inp_total.setPlaceholderText("Auto")
            self._inp_total.setValidator(num_validator)
            self._inp_total.textChanged.connect(self._on_total_changed)
            self._inp_total.returnPressed.connect(self._add_inline_transaction)
            self.table.setCellWidget(0, 6, self._inp_total)

            btn_add = QPushButton("+ Add")
            btn_add.setObjectName("btn_save")
            btn_add.setStyleSheet("QPushButton { background-color: #6366f1; color: white; border-radius: 6px; padding: 6px 14px; font-size: 13px; font-weight: 600; border: none; } QPushButton:hover { background-color: #4f46e5; }")
            btn_add.clicked.connect(self._add_inline_transaction)
            self.table.setCellWidget(0, 7, btn_add)

        offset = 1 if show_input else 0
        if not show_input:
            self.table.setRowCount(len(clean) + 1)  # data + total
        for i in range(len(clean)):
            table_row = i + offset
            row_status = str(clean.iloc[i]['Status']) if 'Status' in clean.columns else ''
            row_bg = None
            if row_status == 'To Receive':
                row_bg = QColor('#f0fdf4')
            elif row_status == 'To Pay':
                row_bg = QColor('#fef2f2')
            for j, col in enumerate(cols):
                val = clean.iloc[i][col]
                s = str(val) if pd.notna(val) else ''
                if s.lower() in ('nan', 'none', 'nat'):
                    s = ''
                if j in (5, 6):
                    try:
                        s = indian_format(float(val))
                    except (ValueError, TypeError):
                        pass
                item = QTableWidgetItem(s)
                if j == 1:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if meta is not None:
                    item.setData(Qt.ItemDataRole.UserRole, (meta.iloc[i]['_sheet_name'], meta.iloc[i]['_sheet_row']))
                if j in (5, 6):
                    try:
                        item.setData(Qt.ItemDataRole.UserRole + 1, float(val))
                    except (ValueError, TypeError):
                        pass
                if j == 3:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if j == 1 and self._current_party is not None:
                    item.setBackground(QColor('#eff6ff'))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                elif row_bg is not None:
                    item.setBackground(row_bg)
                self.table.setItem(table_row, j, item)
            btn = QPushButton("✕")
            btn.setObjectName("btn_remove")
            btn.setFixedSize(28, 28)
            btn.clicked.connect(lambda checked, r=table_row: self.remove_transaction(r))
            self.table.setCellWidget(table_row, 7, btn)
            self.table.resizeRowToContents(table_row)
        
        # Grand total row
        total_row = offset + len(clean)
        self.table.setRowHeight(total_row, 48)
        gt_item = QTableWidgetItem("  Grand Total")
        gt_font = gt_item.font()
        gt_font.setBold(True)
        gt_font.setPointSize(11)
        gt_item.setFont(gt_font)
        gt_item.setBackground(QColor('#f8fafc'))
        gt_item.setForeground(QColor('#1e293b'))
        gt_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(total_row, 0, gt_item)
        for j in range(1, 6):
            item = QTableWidgetItem("")
            item.setBackground(QColor('#f1f5f9'))
            self.table.setItem(total_row, j, item)
        total_val = QTableWidgetItem(indian_format(grand_total))
        total_val.setFont(gt_font)
        total_val.setBackground(QColor('#f1f5f9'))
        total_val.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(total_row, 6, total_val)
        
        # Update party list
        self.party_list.blockSignals(True)
        self.party_list.clear()
        xl = pd.ExcelFile(self.file_path)
        parties = [sheet for sheet in xl.sheet_names if sheet != 'Sheet1']
        for name in parties:
            item = QListWidgetItem(name)
            item.setData(Qt.ItemDataRole.UserRole, name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.party_list.addItem(item)
        self.party_list.blockSignals(False)
        # Update header title and party count
        if self._current_party:
            self._header_title.setText(self._current_party)
        else:
            self._header_title.setText("All Transactions")
        self.party_count_label.setText(f"{len(parties)} party{'ies' if len(parties) != 1 else 'y'}")
        # Re-select current filtered party
        if self._current_party is not None:
            items = self.party_list.findItems(self._current_party, Qt.MatchFlag.MatchExactly)
            if items:
                self.party_list.setCurrentItem(items[0])
        self.table.blockSignals(False)

    def _on_cell_changed(self, row, col):
        item = self.table.item(row, 0)
        if not item:
            return
        meta = item.data(Qt.ItemDataRole.UserRole)
        if not meta:
            return
        sheet_name, sheet_row = meta
        col_map = {0: 'Date', 2: 'Description', 4: 'Quantity', 5: 'Rate', 6: 'Total'}
        if col not in col_map:
            return
        col_name = col_map[col]
        new_val = self.table.item(row, col).text().strip()
        if col in (4, 5, 6):
            new_val = new_val.replace(',', '')
            try:
                new_val = float(new_val)
            except ValueError:
                return
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        if sheet_row >= len(df):
            return
        df.loc[sheet_row, col_name] = new_val
        if col in (4, 5):
            other_col = 5 if col == 4 else 4
            other_item = self.table.item(row, other_col)
            if other_item:
                other_val = other_item.text().strip().replace(',', '')
                try:
                    other_val = float(other_val)
                    total = new_val * other_val
                    df.loc[sheet_row, 'Total'] = total
                except ValueError:
                    pass
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        if self._current_party:
            all_df = self.load_all_data()
            self.refresh_view(all_df[all_df['Party Name'] == self._current_party])
        else:
            self.refresh_view()

    def _on_cell_double_clicked(self, row, col):
        if col != 3:
            return
        item = self.table.item(row, 0)
        if not item:
            return
        meta = item.data(Qt.ItemDataRole.UserRole)
        if not meta:
            return
        sheet_name, sheet_row = meta
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        if sheet_row >= len(df):
            return
        old_status = str(df.loc[sheet_row, 'Status']) if 'Status' in df.columns else 'Credit'
        new_status = 'To Pay' if old_status == 'To Receive' else 'To Receive'
        self._update_status(row, new_status)

    def print_report(self):
        party_name = getattr(self, '_current_party', None)
        df = self.load_all_data()
        if party_name:
            df = df[df['Party Name'] == party_name]
            title = f"Party Ledger – {party_name}"
            suggested = f"{party_name}_ledger.pdf"
        else:
            title = "Party Ledger – All Parties"
            suggested = "all_parties_ledger.pdf"

        total_till_now = None
        if self._date_from and self._date_to:
            from_dt = pd.to_datetime(self._date_from, format='%d-%m-%Y')
            to_dt = pd.to_datetime(self._date_to, format='%d-%m-%Y')
            # Calculate total before filter start
            before_df = df.copy()
            before_df['_pdf_date'] = pd.to_datetime(before_df['Date'], format='%d-%m-%Y', errors='coerce')
            before_df = before_df[before_df['_pdf_date'] < from_dt]
            if not before_df.empty:
                before_df['Total'] = pd.to_numeric(before_df.get('Total', 0), errors='coerce').fillna(0)
                calc = pd.to_numeric(before_df['Quantity'], errors='coerce').fillna(0) * pd.to_numeric(before_df['Rate'], errors='coerce').fillna(0)
                mask = (before_df['Total'] == 0) & (calc != 0)
                before_df.loc[mask, 'Total'] = calc[mask]
                credit_mask = before_df['Status'] == 'To Receive'
                ttn = before_df.loc[credit_mask, 'Total'].sum() - before_df.loc[~credit_mask, 'Total'].sum()
                status = 'To Receive' if ttn >= 0 else 'To Pay'
                total_till_now = (status, abs(ttn))
            # Now apply date filter
            df['_pdf_date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y', errors='coerce')
            df = df[(df['_pdf_date'] >= from_dt) & (df['_pdf_date'] <= to_dt)]
            df = df.drop(columns=['_pdf_date'])

        if df.empty and not total_till_now:
            QMessageBox.information(self, "Export PDF", "No transactions to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF Report", suggested,
            "PDF Files (*.pdf)")
        if not file_path:
            return

        df = df.copy()
        df['Total'] = pd.to_numeric(df.get('Total', 0), errors='coerce').fillna(0)
        calc = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0) * pd.to_numeric(df['Rate'], errors='coerce').fillna(0)
        mask = (df['Total'] == 0) & (calc != 0)
        df.loc[mask, 'Total'] = calc[mask]
        df['_date_parsed'] = pd.to_datetime(df['Date'], format='%d-%m-%Y', errors='coerce')
        df = df.sort_values(by='_date_parsed', ascending=False)

        if 'Status' not in df.columns:
            df['Status'] = 'Credit'
        credit_mask = df['Status'] == 'To Receive'
        grand_total = df.loc[credit_mask, 'Total'].sum() - df.loc[~credit_mask, 'Total'].sum()
        if total_till_now:
            grand_total += total_till_now[1] if total_till_now[0] == 'To Receive' else -total_till_now[1]
        show_party = party_name is None

        cols = ['Date', 'Description', 'Status', 'Quantity', 'Rate', 'Total']
        headers = ['Date', 'Description', 'Status', 'Qty', 'Rate', 'Total']
        if show_party:
            cols = ['Date', 'Party Name', 'Description', 'Status', 'Quantity', 'Rate', 'Total']
            headers = ['Date', 'Party', 'Description', 'Status', 'Qty', 'Rate', 'Total']

        rows_html = ''
        if total_till_now:
            status_label, ttn_amount = total_till_now
            ttn_row = {'Date': '', 'Description': 'Total Till Now', 'Status': status_label, 'Quantity': '', 'Rate': '', 'Total': indian_format(ttn_amount)}
            if show_party:
                ttn_row['Party Name'] = ''
            rows_html += '<tr style="font-weight: bold; background-color: #fef9c3;">'
            for col in cols:
                rows_html += f'<td>{ttn_row.get(col, "")}</td>'
            rows_html += '</tr>'
        for _, row in df.iterrows():
            rows_html += '<tr>'
            for col in cols:
                val = row.get(col, '')
                s = str(val) if pd.notna(val) else ''
                if s.lower() in ('nan', 'none', 'nat'):
                    s = ''
                if col in ('Rate', 'Total'):
                    try:
                        s = indian_format(float(val))
                    except (ValueError, TypeError):
                        pass
                rows_html += f'<td>{s}</td>'
            rows_html += '</tr>'
        n_cols = len(cols)
        colspan = n_cols - 1
        gt_label = "Overall Grand Total" if total_till_now else "Grand Total"
        rows_html += f'<tr style="font-weight: bold; background-color: #f1f5f9;"><td colspan="{colspan}" style="text-align: right; padding: 8px;">{gt_label}</td><td style="padding: 8px;">{indian_format(grand_total)}</td></tr>'

        today = date.today().strftime('%d-%m-%Y %H:%M')
        header_cells = ''.join(f'<th>{h}</th>' for h in headers)
        html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 11pt; }}
  h1 {{ font-size: 20pt; color: #1e293b; margin-bottom: 6pt; }}
  .meta {{ color: #64748b; font-size: 10pt; margin-bottom: 12pt; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background-color: #e2e8f0; font-weight: bold; padding: 8px; border: 1px solid #64748b; }}
  td {{ padding: 6px; border: 1px solid #64748b; }}
  .footer {{ margin-top: 15pt; font-size: 9pt; color: #94a3b8; text-align: center; }}
</style></head><body>
<h1>{title}</h1>
<div class="meta">Exported: {today} &nbsp;|&nbsp; Transactions: {len(df)}</div>
<table border="1" cellspacing="0" cellpadding="5" width="100%" style="border-collapse: collapse;">
  <tr bgcolor="#e2e8f0">
    {header_cells}
  </tr>
  {rows_html}
</table>
<div class="footer">Generated by Party Ledger Manager</div>
</body></html>'''

        try:
            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
            doc.print(printer)

            QMessageBox.information(self, "Export PDF",
                f"PDF saved successfully:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export PDF Error",
                f"Failed to create PDF:\n{str(e)}")

    def _apply_date_filter(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Date Filter")
        dialog.resize(400, 260)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(16)
        layout.setContentsMargins(28, 28, 28, 28)

        heading = QLabel("Filter by Date Range")
        heading.setStyleSheet("font-size: 18px; font-weight: 700; color: #0f172a;")
        layout.addWidget(heading)

        from_input = QLineEdit()
        from_input.setInputMask("00-00-0000;_")
        from_input.setPlaceholderText("DD-MM-YYYY")
        if self._date_from:
            from_input.setText(self._date_from)

        to_input = QLineEdit()
        to_input.setInputMask("00-00-0000;_")
        to_input.setPlaceholderText("DD-MM-YYYY")
        if self._date_to:
            to_input.setText(self._date_to)

        form = QFormLayout()
        form.setSpacing(12)
        form.addRow("From:", from_input)
        form.addRow("To:", to_input)
        layout.addLayout(form)

        btn_box = QDialogButtonBox()
        btn_box.addButton(QDialogButtonBox.StandardButton.Ok)
        btn_box.addButton(QDialogButtonBox.StandardButton.Cancel)
        clear_btn = btn_box.addButton("Clear Filter", QDialogButtonBox.ButtonRole.ResetRole)
        clear_btn.setStyleSheet("QPushButton { background-color: #ffffff; color: #ef4444; border: 1px solid #fecaca; border-radius: 8px; padding: 8px 16px; font-size: 13px; font-weight: 500; } QPushButton:hover { background-color: #fef2f2; border-color: #f87171; } QPushButton:pressed { background-color: #fee2e2; }")
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)

        cleared = False
        def do_clear():
            nonlocal cleared
            cleared = True
            dialog.accept()
        clear_btn.clicked.connect(do_clear)

        layout.addWidget(btn_box)

        if dialog.exec():
            if cleared:
                self._date_from = None
                self._date_to = None
                self.btn_date_filter.setText("  Date Filter")
                self.btn_date_filter.setObjectName("btn_date_filter")
                self.btn_date_filter.style().unpolish(self.btn_date_filter)
                self.btn_date_filter.style().polish(self.btn_date_filter)
                if self._current_party:
                    all_df = self.load_all_data()
                    self.refresh_view(all_df[all_df['Party Name'] == self._current_party])
                else:
                    self.refresh_view()
            else:
                f = from_input.text().strip()
                t = to_input.text().strip()
                if not f or not t:
                    QMessageBox.warning(dialog, "Date Filter", "Enter both From and To dates in DD-MM-YYYY format, or use Clear Filter.")
                    return
                fd = QDate.fromString(f, "dd-MM-yyyy")
                td = QDate.fromString(t, "dd-MM-yyyy")
                if not fd.isValid() or not td.isValid():
                    QMessageBox.warning(dialog, "Date Filter", "Invalid date format. Use DD-MM-YYYY (e.g. 01-06-2026).")
                    return
                if fd > td:
                    f, t = t, f
                self._date_from = f
                self._date_to = t
                self.btn_date_filter.setText(f"  Filter: {f} — {t}")
                self.btn_date_filter.setObjectName("btn_active_filter")
                self.btn_date_filter.style().unpolish(self.btn_date_filter)
                self.btn_date_filter.style().polish(self.btn_date_filter)
                if self._current_party:
                    all_df = self.load_all_data()
                    self.refresh_view(all_df[all_df['Party Name'] == self._current_party])
                else:
                    self.refresh_view()

    def _set_status_filter(self, status):
        self._status_filter = status
        # Update button checked states
        self._btn_filter_all.setChecked(status is None)
        self._btn_filter_receive.setChecked(status == 'To Receive')
        self._btn_filter_pay.setChecked(status == 'To Pay')
        # Re-load data and re-apply current party filter if any
        df = self.load_all_data()
        if self._current_party:
            df = df[df['Party Name'] == self._current_party]
        self.refresh_view(df)

    def show_party_summary(self):
        dialog = PartySummaryDialog(self.file_path, self)
        dialog.exec()

    def filter_party_list(self, text):
        for i in range(self.party_list.count()):
            item = self.party_list.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def filter_by_party(self, item):
        party_name = item.text()
        # Toggle: click same party again shows all
        if hasattr(self, '_current_party') and self._current_party == party_name:
            self._current_party = None
            self.refresh_view()
            return
        self._current_party = party_name
        df = self.load_all_data()
        filtered_df = df[df['Party Name'] == party_name]
        self.refresh_view(filtered_df)

    def add_party(self):
        dialog = CreatePartyDialog(self)
        if dialog.exec():
            name = dialog.get_name()
            with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                pd.DataFrame(columns=self.SHEET_COLUMNS).to_excel(writer, sheet_name=name, index=False)
            self.refresh_view()

    def add_transaction(self):
        xl = pd.ExcelFile(self.file_path)
        parties = [sheet for sheet in xl.sheet_names if sheet != 'Sheet1']
        dialog = TransactionDialog(parties, self)
        if dialog.exec():
            data = dialog.get_data()
            if data:
                try:
                    df = pd.read_excel(self.file_path, sheet_name=data['Party Name'])
                except Exception:
                    QMessageBox.warning(self, "Error", f"Party '{data['Party Name']}' not found.")
                    return
                # Ensure all expected columns exist
                for col in self.SHEET_COLUMNS:
                    if col not in df.columns:
                        df[col] = 0.0
                df = df[self.SHEET_COLUMNS]
                
                new_row = pd.DataFrame([{
                    'Date': data['Date'],
                    'Description': data['Description'],
                    'Quantity': data['Quantity'],
                    'Rate': data['Rate'],
                    'Status': data.get('Status', 'Credit'),
                    'Total': data['Total']
                }])
                updated_df = pd.concat([df, new_row], ignore_index=True)
                
                with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    updated_df.to_excel(writer, sheet_name=data['Party Name'], index=False)
                self.refresh_view()
            else:
                QMessageBox.warning(self, "Error", "Invalid data input")

    def eventFilter(self, obj, event):
        if obj is getattr(self, '_inp_desc', None) and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace) and self._current_suggestion:
                completer = getattr(self, '_desc_completer', None)
                if completer and completer.popup().isVisible():
                    desc = self._current_suggestion
                    reply = QMessageBox.question(self, "Remove Suggestion",
                        f"Remove \"{desc}\" from suggestions permanently?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                    if reply == QMessageBox.StandardButton.Yes:
                        self._remove_desc_suggestion(desc)
                    return True
        if obj is getattr(self, '_desc_popup_viewport', None) and event.type() == QEvent.Type.MouseButtonPress:
            if event.button() == Qt.MouseButton.RightButton:
                index = obj.parent().indexAt(event.pos())
                if index.isValid():
                    desc = index.data()
                    menu = QMenu()
                    remove_action = menu.addAction(f"Remove \"{desc}\" from suggestions")
                    action = menu.exec(event.globalPosition().toPoint())
                    if action == remove_action:
                        self._remove_desc_suggestion(desc)
                    return True
        return super().eventFilter(obj, event)

    def _on_suggestion_highlighted(self, text):
        self._current_suggestion = text

    def _on_desc_selected(self, text):
        if not self._current_party:
            return
        try:
            df = pd.read_excel(self.file_path, sheet_name=self._current_party)
            if df.empty:
                return
            matches = df[df['Description'].astype(str).str.strip() == text.strip()]
            if matches.empty:
                return
            last_rate = pd.to_numeric(matches['Rate'].iloc[-1], errors='coerce')
            if pd.notna(last_rate) and last_rate > 0:
                self._inp_rate.setText(str(last_rate))
        except Exception:
            pass

    def _update_status(self, row, new_status):
        item = self.table.item(row, 0)
        if not item:
            return
        meta = item.data(Qt.ItemDataRole.UserRole)
        if not meta:
            return
        sheet_name, sheet_row = meta
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        if sheet_row >= len(df):
            return
        old_status = str(df.loc[sheet_row, 'Status']) if 'Status' in df.columns else 'Credit'

        parts = []
        for c in range(7):
            it = self.table.item(row, c)
            if it:
                label = ['Date', 'Party', 'Description', 'Status', 'Qty', 'Rate', 'Total'][c]
                parts.append(f"{label}: {it.text()}")
        context = '\n'.join(parts)

        reply = QMessageBox.question(self, "Change Status",
            f"Change status from {old_status} to {new_status}?\n\n{context}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return

        df.loc[sheet_row, 'Status'] = new_status
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        if self._current_party:
            all_df = self.load_all_data()
            self.refresh_view(all_df[all_df['Party Name'] == self._current_party])
        else:
            self.refresh_view()

    def remove_transaction(self, row):
        item = self.table.item(row, 0)
        if not item:
            return
        meta = item.data(Qt.ItemDataRole.UserRole)
        if not meta:
            return
        sheet_name, sheet_row = meta
        reply = QMessageBox.question(self, "Remove Transaction",
            f"Delete this transaction from '{sheet_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        df = df.drop(index=sheet_row, errors='ignore').reset_index(drop=True)
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        if self._current_party:
            df = self.load_all_data()
            self.refresh_view(df[df['Party Name'] == self._current_party])
        else:
            self.refresh_view()

    def _rename_party(self, item):
        old_name = item.data(Qt.ItemDataRole.UserRole)
        new_name = item.text().strip()
        if not new_name or old_name == new_name:
            self.party_list.blockSignals(True)
            item.setText(old_name)
            self.party_list.blockSignals(False)
            return
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path)
        if new_name in wb.sheetnames:
            QMessageBox.warning(self, "Rename Party", f"Party \"{new_name}\" already exists.")
            self.party_list.blockSignals(True)
            item.setText(old_name)
            self.party_list.blockSignals(False)
            wb.close()
            return
        ws = wb[old_name]
        ws.title = new_name
        wb.save(self.file_path)
        wb.close()
        item.setData(Qt.ItemDataRole.UserRole, new_name)
        if self._current_party == old_name:
            self._current_party = new_name
        if self._current_party:
            self._header_title.setText(self._current_party)

    def remove_party(self):
        selected = self.party_list.currentItem()
        if not selected:
            QMessageBox.warning(self, "Remove Party", "Select a party from the list first.")
            return
        party_name = selected.text()
        reply = QMessageBox.question(self, "Remove Party",
            f"Delete party '{party_name}' and all its transactions?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            pd.DataFrame().to_excel(writer, sheet_name=party_name, index=False)
        # Delete the sheet entirely
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path)
        if party_name in wb.sheetnames:
            del wb[party_name]
        wb.save(self.file_path)
        wb.close()
        if self._current_party == party_name:
            self._current_party = None
        self.refresh_view()

    def _on_total_changed(self):
        total_str = self._inp_total.text().strip()
        if total_str:
            self._inp_qty.blockSignals(True)
            self._inp_rate.blockSignals(True)
            self._inp_qty.clear()
            self._inp_rate.clear()
            self._inp_qty.setReadOnly(True)
            self._inp_rate.setReadOnly(True)
            self._inp_qty.blockSignals(False)
            self._inp_rate.blockSignals(False)
        else:
            self._inp_qty.setReadOnly(False)
            self._inp_rate.setReadOnly(False)

    def _inline_auto_calc(self):
        qty_str = self._inp_qty.text().strip()
        rate_str = self._inp_rate.text().strip()
        if qty_str and rate_str:
            try:
                qty = float(qty_str)
                rate = float(rate_str)
                self._inp_total.blockSignals(True)
                self._inp_total.setReadOnly(True)
                self._inp_total.setText(str(qty * rate))
                self._inp_total.blockSignals(False)
            except ValueError:
                pass
        elif qty_str or rate_str:
            if not self._inp_total.isReadOnly():
                self._inp_total.blockSignals(True)
                self._inp_total.clear()
                self._inp_total.setReadOnly(True)
                self._inp_total.blockSignals(False)
        else:
            if self._inp_total.isReadOnly():
                self._inp_total.blockSignals(True)
                self._inp_total.clear()
                self._inp_total.setReadOnly(False)
                self._inp_total.blockSignals(False)

    def _add_inline_transaction(self):
        date_str = self._inp_date.text().strip()
        if not self._inp_date.hasAcceptableInput():
            QMessageBox.warning(self, "Validation Error", "Enter a complete date (DD-MM-YYYY).")
            return
        desc_str = self._inp_desc.text().strip()
        qty_str = self._inp_qty.text().strip()
        rate_str = self._inp_rate.text().strip()
        total_str = self._inp_total.text().strip()
        party_name = self._current_party

        if not date_str:
            QMessageBox.warning(self, "Validation Error", "Date is required.")
            return
        if not party_name:
            QMessageBox.warning(self, "Validation Error", "No party selected.")
            return

        qty = float(qty_str) if qty_str else 0.0
        rate = float(rate_str) if rate_str else 0.0
        total = float(total_str) if total_str else (qty * rate)

        if not ((qty_str and rate_str) or total_str):
            QMessageBox.warning(self, "Validation Error",
                "Enter both Quantity & Rate, or enter Total directly.")
            return

        try:
            df = pd.read_excel(self.file_path, sheet_name=party_name)
        except Exception:
            QMessageBox.warning(self, "Error", f"Party '{party_name}' not found.")
            return
        for col in self.SHEET_COLUMNS:
            if col not in df.columns:
                df[col] = 0.0
        df = df[self.SHEET_COLUMNS]

        new_row = pd.DataFrame([{
            'Date': date_str, 'Description': desc_str,
            'Quantity': qty, 'Rate': rate,
            'Status': self._inp_status.currentText(),
            'Total': total
        }])
        updated_df = pd.concat([df, new_row], ignore_index=True)

        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_df.to_excel(writer, sheet_name=party_name, index=False)
        df = self.load_all_data()
        self.refresh_view(df[df['Party Name'] == party_name])

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = LedgerApp()
    window.show()
    sys.exit(app.exec())
